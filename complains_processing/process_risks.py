#!/usr/bin/env python3
"""
Скрипт для обработки комментариев через Gemini API и определения рисков.
Обрабатывает CSV файл построчно и заполняет категории, подкатегории и уровни рисков.
"""

import csv
import json
import os
import time
from google import genai
from google.genai import types
from pathlib import Path

# Попытка импортировать openpyxl для Excel (опционально)
try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# API ключ для Gemini
GEMINI_API_KEY = "AIzaSyCnJmahkTySYiRQF0P89iIac0gIWYUrf7s"

# Инициализация клиента Gemini (создается один раз)
_gemini_client = None

def get_gemini_client():
    """
    Получает инициализированный клиент Gemini.
    """
    global _gemini_client
    if _gemini_client is not None:
        return _gemini_client
    
    # Создаем клиент с API ключом
    _gemini_client = genai.Client(api_key=GEMINI_API_KEY)
    print("✅ Клиент Gemini инициализирован")
    return _gemini_client

# Допустимые категории и подкатегории (для валидации)
VALID_CATEGORIES = {
    "Operational Risk": ["Technical Failure", "Transaction Issue", "Access & Identity", "System Downtime", "Technical Support"],
    "Reputational Risk": ["Customer Service", "Ethical & Trust", "Fee Transparency"],
    "Liquidity Risk": ["Withdrawal Limits", "Market Panic", "Currency Availability"],
    "Security & Fraud": ["Phishing & Scams", "Account Takeover", "Data Privacy", "Safety"],
    "Compliance & Legal": ["Account Freezing", "Regulatory/Sharia"],
    "General": ["Neutral", "Spam/Neutral", "Feedback", "Neutral (Competitor)"]
}

VALID_RISK_LEVELS = ["Low", "Medium", "High", "Critical"]

# Системные инструкции для модели (загружаются один раз)
SYSTEM_INSTRUCTIONS = """I'm compiling a reference guide for categorizing social media comments about a bank in Somalia. I need your help to categorize each comment.

I will provide you with social media comments that include Somali text and English translations. Please help me categorize them according to the structure below.

RISK CATEGORIES AND SUBCATEGORIES:

1. Operational Risk
   - Technical Failure: App crashes, UI/UX bugs, website login errors, server-side issues
   - Transaction Issue: Money debited but not received, remittance delays, failed transfers
   - Access & Identity: OTP (SMS) not arriving, blocked passwords, login authentication failures
   - System Downtime: "The whole system is down", "Bank is offline", widespread outages
   - Technical Support: Questions about how to use features, account setup, technical help requests

2. Reputational Risk
   - Customer Service: Rude staff, ignored support tickets, long wait times on phone or in-branch
   - Ethical & Trust: Allegations of corruption, unfair treatment, rumors of insolvency or scams
   - Fee Transparency: Complaints about "hidden" charges, unexplained commissions, high rates

3. Liquidity Risk
   - Withdrawal Limits: Unable to withdraw cash from ATMs or branches, daily limit restrictions
   - Market Panic: "Run on the bank" signals: "Take your money out now before they close!"
   - Currency Availability: Shortage of USD or local currency (common in the Somali context)

4. Security & Fraud
   - Phishing & Scams: Reports of fake WhatsApp groups, fraudulent SMS, or fake bank pages
   - Account Takeover: "My account was hacked", "Money disappeared without my knowledge"
   - Data Privacy: Concerns over leaked personal info or bank statements shared publicly
   - Safety: Concerns about physical safety, theft, security of transactions or cards

5. Compliance & Legal
   - Account Freezing: Accounts blocked due to AML/KYC checks, "Bank won't release my funds"
   - Regulatory/Sharia: Complaints regarding non-compliance with Islamic banking or local laws

6. General
   - Neutral: General positive or neutral comments, compliments, general questions
   - Spam/Neutral: Spam messages, irrelevant content, promotional content
   - Feedback: Feature requests, suggestions, general feedback without risk
   - Neutral (Competitor): Mentions of competitors without negative sentiment

RISK LEVELS:
- Low: Routine inquiries or minor dissatisfaction with no direct threat to the bank. Examples: General questions about branch hours, exchange rates, or simple feature requests.
- Medium: Individual service issues or isolated technical bugs. Examples: "I forgot my password", "The app is slow today", or complaints about a specific teller.
- High: Serious financial or technical issues affecting trust or money. Examples: Failed transactions, money missing from account, or allegations of fraud/scams.
- Critical: Systemic threats that could cause mass panic or widespread failure. Examples: App-wide outages, "Bank Run" calls (withdraw everything!), or confirmed security breaches.

Please help me categorize each comment by:
- Using EXACT category and subcategory names as listed above
- Matching the subcategory to the category correctly
- Choosing the most appropriate risk level based on severity
- If comment is neutral/spam/feedback without risk, use General category with appropriate subcategory and Low level

Please return your categorization in JSON format only, with these exact keys:
{
  "risk_category": "exact category name from above",
  "risk_subcategory": "exact subcategory name from above",
  "risk_level": "Low/Medium/High/Critical"
}

Return ONLY valid JSON, no other text."""

def analyze_risk_local(somali_text, english_text):
    """
    Анализирует комментарий локально (без API) и определяет категории рисков.
    Использует паттерны ключевых слов и контекстный анализ.
    """
    # Объединяем тексты для анализа
    combined_text = f"{somali_text} {english_text}".lower()
    
    # Паттерны для определения категорий и подкатегорий
    patterns = {
        "Operational Risk": {
            "Technical Failure": ["crash", "bug", "error", "not working", "broken", "fail", "doesn't work", "not functioning"],
            "Transaction Issue": ["transaction", "debit", "credit", "transfer", "remittance", "money not received", "payment failed", "deducted", "withdrawal"],
            "Access & Identity": ["password", "login", "otp", "sms", "authentication", "blocked", "access", "account locked"],
            "System Downtime": ["system down", "offline", "server", "outage", "not available", "down"],
            "Technical Support": ["how to", "how can", "help", "support", "question", "setup", "create account", "open account"]
        },
        "Reputational Risk": {
            "Customer Service": ["rude", "ignored", "wait", "slow", "bad service", "poor service", "help me", "customer care", "support"],
            "Ethical & Trust": ["corruption", "scam", "fraud", "thief", "steal", "unfair", "dishonest", "trust", "not trustworthy"],
            "Fee Transparency": ["fee", "charge", "commission", "cost", "hidden", "unexplained", "expensive", "high rate", "deduct"]
        },
        "Liquidity Risk": {
            "Withdrawal Limits": ["withdraw", "withdrawal", "limit", "cash", "atm", "can't withdraw"],
            "Market Panic": ["take your money", "withdraw now", "close", "bank run", "panic", "get out"],
            "Currency Availability": ["currency", "usd", "dollar", "shortage", "not available"]
        },
        "Security & Fraud": {
            "Phishing & Scams": ["phishing", "fake", "scam", "fraudulent", "fraud"],
            "Account Takeover": ["hacked", "hack", "stolen", "disappeared", "missing money", "unauthorized"],
            "Data Privacy": ["privacy", "leaked", "data", "personal info", "information", "statement"],
            "Safety": ["safety", "safe", "secure", "security", "theft", "steal", "rob", "afraid", "fear", "danger"]
        },
        "Compliance & Legal": {
            "Account Freezing": ["frozen", "blocked", "freeze", "can't access", "won't release", "aml", "kyc"],
            "Regulatory/Sharia": ["sharia", "islamic", "halal", "haram", "regulatory", "law", "compliance", "shirk"]
        },
        "General": {
            "Neutral": ["good", "great", "excellent", "thank", "thanks", "pray", "prayer", "compliment"],
            "Spam/Neutral": ["send me", "money", "dollar", "please send", "help me with money"],
            "Feedback": ["idea", "suggestion", "feature", "improve", "better", "recommend"],
            "Neutral (Competitor)": ["waafi", "dahabshiil", "salaam bank", "competitor", "other bank"]
        }
    }
    
    # Определяем уровень риска на основе ключевых слов
    high_risk_keywords = ["hacked", "stolen", "fraud", "scam", "corruption", "thief", "steal", "missing money", "failed transaction", "money disappeared"]
    critical_keywords = ["bank run", "withdraw everything", "close", "panic", "system down", "outage", "everyone"]
    medium_risk_keywords = ["problem", "issue", "bad", "poor", "slow", "error", "not working", "complaint"]
    
    # Подсчитываем совпадения для каждой категории
    category_scores = {}
    subcategory_scores = {}
    
    for category, subcategories in patterns.items():
        category_score = 0
        for subcategory, keywords in subcategories.items():
            score = sum(1 for keyword in keywords if keyword in combined_text)
            if score > 0:
                subcategory_scores[f"{category} > {subcategory}"] = score
                category_score += score
        if category_score > 0:
            category_scores[category] = category_score
    
    # Определяем категорию и подкатегорию
    if not category_scores:
        # Если нет совпадений, проверяем на нейтральный контент
        if any(word in combined_text for word in ["good", "great", "thank", "pray", "compliment"]):
            return {
                "risk_category": "General",
                "risk_subcategory": "Neutral",
                "risk_level": "Low"
            }
        else:
            return {
                "risk_category": "General",
                "risk_subcategory": "Neutral",
                "risk_level": "Low"
            }
    
    # Выбираем категорию с наибольшим счетом
    best_category = max(category_scores, key=lambda k: category_scores[k])
    
    # Выбираем подкатегорию для этой категории
    best_subcategory = None
    best_subcategory_score = 0
    
    for key, score in subcategory_scores.items():
        if key.startswith(best_category + " >"):
            if score > best_subcategory_score:
                best_subcategory_score = score
                best_subcategory = key.split(" > ")[1]
    
    if not best_subcategory:
        # Если не нашли подкатегорию, используем первую доступную
        best_subcategory = VALID_CATEGORIES.get(best_category, ["Neutral"])[0]
    
    # Определяем уровень риска
    risk_level = "Low"
    if any(keyword in combined_text for keyword in critical_keywords):
        risk_level = "Critical"
    elif any(keyword in combined_text for keyword in high_risk_keywords):
        risk_level = "High"
    elif any(keyword in combined_text for keyword in medium_risk_keywords):
        risk_level = "Medium"
    
    return {
        "risk_category": best_category,
        "risk_subcategory": best_subcategory,
        "risk_level": risk_level
    }


def get_risk_assessment(somali_text, english_text, use_api=False):
    """
    Анализирует комментарий и определяет категории рисков.
    Может работать через API или локально (без API).
    
    Args:
        somali_text: Текст комментария на сомалийском
        english_text: Перевод комментария на английский
        use_api: Если True - использует Gemini API, если False - локальный анализ
    """
    # Если не используем API, используем локальный анализ
    if not use_api:
        print(f"  🤖 Локальный анализ (без API)")
        return analyze_risk_local(somali_text, english_text)
    
    # Получаем инициализированный клиент
    client = get_gemini_client()
    
    # Формируем краткий промпт только с данными комментария
    # Все инструкции передаются через system_instruction в config
    prompt = f"""Please help me categorize this comment for my reference guide:

Somali text: {somali_text}
English translation: {english_text}

Please categorize it according to the structure I provided and return JSON format only."""

    # Выводим промпт в консоль
    print("\n" + "=" * 80)
    print("ПРОМПТ ОТПРАВЛЯЕМЫЙ В GEMINI:")
    print("=" * 80)
    print(prompt)
    print("=" * 80 + "\n")

    try:
        # Настройки безопасности - делаем менее строгими для анализа банковских комментариев
        # Это позволяет обрабатывать комментарии с упоминаниями проблем, угроз и т.д.
        safety_settings = [
            types.SafetySetting(
                category=types.HarmCategory.HARM_CATEGORY_HARASSMENT,
                threshold=types.SafetySetting.HarmBlockThreshold.BLOCK_ONLY_HIGH
            ),
            types.SafetySetting(
                category=types.HarmCategory.HARM_CATEGORY_HATE_SPEECH,
                threshold=types.SafetySetting.HarmBlockThreshold.BLOCK_ONLY_HIGH
            ),
            types.SafetySetting(
                category=types.HarmCategory.HARM_CATEGORY_SEXUALLY_EXPLICIT,
                threshold=types.SafetySetting.HarmBlockThreshold.BLOCK_MEDIUM_AND_ABOVE
            ),
            types.SafetySetting(
                category=types.HarmCategory.HARM_CATEGORY_DANGEROUS_CONTENT,
                threshold=types.SafetySetting.HarmBlockThreshold.BLOCK_ONLY_HIGH  # Важно: позволяем анализировать угрозы
            ),
        ]
        
        # Отправляем запрос с настройками безопасности и системными инструкциями
        response = client.models.generate_content(
            model="gemini-3-flash-preview",
            contents=prompt,
            config=types.GenerateContentConfig(
                system_instruction=SYSTEM_INSTRUCTIONS,
                safety_settings=safety_settings
            )
        )
        
        # Обработка случая, когда ответ заблокирован из-за безопасности
        if not response.candidates:
            raise Exception("Ответ не содержит кандидатов")
        
        candidate = response.candidates[0]
        finish_reason = candidate.finish_reason
        
        # Проверяем finish_reason (может быть строкой или enum)
        finish_reason_str = str(finish_reason) if finish_reason else None
        
        if finish_reason_str and 'SAFETY' in finish_reason_str.upper():
            print(f"  ⚠️  Ответ заблокирован из-за политики безопасности (SAFETY)")
            print(f"     Комментарий содержит контент, который модель считает потенциально вредным.")
            print(f"     Это может быть из-за упоминаний угроз, насилия, мошенничества и т.д.")
            print(f"     Рекомендуется: ручная проверка комментария")
            
            # Выводим детали ответа от API
            print(f"\n  📋 Детали ответа от API:")
            print(f"     finish_reason: {finish_reason}")
            
            # Выводим информацию о безопасности, если доступна
            if hasattr(candidate, 'safety_ratings') and candidate.safety_ratings:
                print(f"\n     safety_ratings:")
                for rating in candidate.safety_ratings:
                    category = getattr(rating, 'category', 'Unknown')
                    probability = getattr(rating, 'probability', 'Unknown')
                    blocked = getattr(rating, 'blocked', False)
                    print(f"       - {category}: probability={probability}, blocked={blocked}")
            
            # Помечаем как требующий ручной проверки
            return {
                "risk_category": "General",
                "risk_subcategory": "Neutral",
                "risk_level": "High"  # Высокий уровень, так как требует внимания
            }
        elif finish_reason_str and 'STOP' not in finish_reason_str.upper():
            print(f"  ⚠️  Неожиданный finish_reason: {finish_reason}")
            return {
                "risk_category": "",
                "risk_subcategory": "",
                "risk_level": ""
            }
        
        # Извлекаем текст ответа
        # В новом API можно использовать response.text напрямую
        if hasattr(response, 'text') and response.text:
            response_text = response.text.strip()
        elif candidate.content and candidate.content.parts:
            response_text = candidate.content.parts[0].text.strip()
        else:
            raise Exception("Ответ не содержит текста")
        
        # Пытаемся найти JSON в ответе
        # Удаляем markdown форматирование если есть
        if "```json" in response_text:
            response_text = response_text.split("```json")[1].split("```")[0].strip()
        elif "```" in response_text:
            response_text = response_text.split("```")[1].split("```")[0].strip()
        
        # Пытаемся найти JSON объект в тексте
        start_idx = response_text.find("{")
        end_idx = response_text.rfind("}")
        if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
            response_text = response_text[start_idx:end_idx+1]
        
        # Парсим JSON
        result = json.loads(response_text)
        
        category = result.get("risk_category", "").strip()
        subcategory = result.get("risk_subcategory", "").strip()
        level = result.get("risk_level", "").strip()
        
        # Валидация категории
        if category and category not in VALID_CATEGORIES:
            # Пытаемся найти похожую категорию
            category_lower = category.lower()
            for valid_cat in VALID_CATEGORIES.keys():
                if valid_cat.lower() in category_lower or category_lower in valid_cat.lower():
                    category = valid_cat
                    break
            else:
                print(f"  ⚠️  Неизвестная категория: {category}")
        
        # Валидация подкатегории
        if category and subcategory:
            valid_subs = VALID_CATEGORIES.get(category, [])
            if subcategory not in valid_subs:
                # Пытаемся найти похожую подкатегорию
                subcategory_lower = subcategory.lower()
                for valid_sub in valid_subs:
                    if valid_sub.lower() in subcategory_lower or subcategory_lower in valid_sub.lower():
                        subcategory = valid_sub
                        break
                else:
                    print(f"  ⚠️  Неизвестная подкатегория '{subcategory}' для категории '{category}'")
        
        # Валидация уровня риска
        if level and level not in VALID_RISK_LEVELS:
            level_lower = level.lower()
            for valid_level in VALID_RISK_LEVELS:
                if valid_level.lower() == level_lower:
                    level = valid_level
                    break
            else:
                print(f"  ⚠️  Неизвестный уровень риска: {level}")
        
        return {
            "risk_category": category,
            "risk_subcategory": subcategory,
            "risk_level": level
        }
    
    except json.JSONDecodeError as e:
        print(f"  ⚠️  Ошибка парсинга JSON: {e}")
        response_text_local = locals().get('response_text', '')
        if response_text_local:
            print(f"  Ответ от API: {response_text_local}")
        return {
            "risk_category": "",
            "risk_subcategory": "",
            "risk_level": ""
        }
    except Exception as e:
        error_msg = str(e)
        print(f"  ⚠️  Ошибка при обращении к API: {error_msg}")
        
        # Специальная обработка ошибки 404 (модель не найдена)
        if "404" in error_msg or "not found" in error_msg.lower():
            print(f"  💡 Проверьте доступность модели gemini-3-flash-preview")
        
        # Детальная информация об ошибке только в режиме отладки
        import os
        if os.environ.get('DEBUG', '').lower() == 'true':
            import traceback
            traceback.print_exc()
        
        return {
            "risk_category": "",
            "risk_subcategory": "",
            "risk_level": ""
        }


def process_csv(input_file, output_file, json_output_file, use_api=False):
    """
    Обрабатывает CSV файл построчно и заполняет риски.
    Сохраняет результаты в отдельный JSON файл.
    
    Args:
        input_file: Путь к входному CSV файлу
        output_file: Путь к выходному CSV/Excel файлу (опционально)
        json_output_file: Путь к выходному JSON файлу
        use_api: Если True - использует Gemini API, если False - локальный анализ
    """
    rows = []
    results = []  # Для сохранения в JSON
    
    # Читаем исходный файл
    print(f"Чтение файла: {input_file}")
    with open(input_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        rows = list(reader)
    
    if not rows:
        print("Файл пуст!")
        return
    
    # Определяем заголовки
    headers = rows[0]
    print(f"Заголовки: {headers}")
    
    # Находим индексы колонок
    somali_idx = headers.index('Somali') if 'Somali' in headers else 0
    english_idx = headers.index('English') if 'English' in headers else 1
    category_idx = headers.index('risk_category') if 'risk_category' in headers else 2
    subcategory_idx = headers.index('risk_subcategory') if 'risk_subcategory' in headers else 3
    level_idx = headers.index('risk_level') if 'risk_level' in headers else 4
    
    # Обрабатываем каждую строку
    total_rows = len(rows) - 1  # Минус заголовок
    processed = 0
    skipped = 0
    
    for i, row in enumerate(rows[1:], start=1):
        # Пропускаем пустые строки
        if not row or len(row) <= somali_idx or not row[somali_idx].strip():
            skipped += 1
            continue
        
        # Проверяем, нужно ли обрабатывать (если уже заполнено, пропускаем)
        if len(row) > category_idx and row[category_idx].strip():
            skipped += 1
            if skipped % 50 == 0:
                print(f"Пропущено {skipped} уже заполненных строк...")
            continue
        
        # Расширяем строку до нужной длины
        while len(row) <= level_idx:
            row.append("")
        
        somali_text = row[somali_idx].strip()
        english_text = row[english_idx].strip() if len(row) > english_idx else ""
        
        if not somali_text:
            skipped += 1
            continue
        
        print(f"\n[{i}/{total_rows}] Обработка строки:")
        print(f"  Somali: {somali_text}")
        if english_text:
            print(f"  English: {english_text}")
        
        # Получаем оценку рисков
        risk_assessment = get_risk_assessment(somali_text, english_text, use_api=use_api)
        
        # Заполняем строку (для CSV, если нужно)
        row[category_idx] = risk_assessment["risk_category"]
        row[subcategory_idx] = risk_assessment["risk_subcategory"]
        row[level_idx] = risk_assessment["risk_level"]
        
        # Сохраняем результат для JSON
        result_entry = {
            "row_number": i,
            "somali": somali_text,
            "english": english_text,
            "risk_category": risk_assessment["risk_category"],
            "risk_subcategory": risk_assessment["risk_subcategory"],
            "risk_level": risk_assessment["risk_level"]
        }
        results.append(result_entry)
        
        if risk_assessment["risk_category"]:
            print(f"  ✅ Результат: {risk_assessment['risk_category']} > {risk_assessment['risk_subcategory']} ({risk_assessment['risk_level']})")
        else:
            print(f"  ❌ Не удалось определить риски")
        
        processed += 1
        
        # Сохраняем промежуточный результат в JSON каждые 5 строк
        if processed % 5 == 0:
            save_json_results(results, json_output_file)
            print(f"  💾 Промежуточное сохранение в JSON... ({processed} обработано)")
        
        # Небольшая задержка только при использовании API
        if use_api:
            time.sleep(0.5)
    
    # Финальное сохранение
    print(f"\nСохранение результатов:")
    print(f"  JSON: {json_output_file}")
    save_json_results(results, json_output_file)
    
    # Опционально: сохраняем также в CSV/Excel если указан output_file
    if output_file:
        print(f"  CSV/Excel: {output_file}")
        save_file(rows, output_file)
    
    print(f"\nОбработка завершена!")
    print(f"  Всего строк: {total_rows}")
    print(f"  Обработано: {processed}")
    print(f"  Пропущено: {skipped}")


def save_file(rows, output_file):
    """
    Сохраняет данные в CSV или Excel файл в зависимости от расширения.
    """
    output_path = Path(output_file)
    
    # Определяем формат по расширению
    if output_path.suffix.lower() in ['.xlsx', '.xls']:
        if not EXCEL_AVAILABLE:
            print("⚠️  openpyxl не установлен. Сохраняю в CSV формат вместо Excel.")
            # Меняем расширение на .csv
            output_file = str(output_path.with_suffix('.csv'))
            save_csv(rows, output_file)
        else:
            save_excel(rows, output_file)
    else:
        save_csv(rows, output_file)


def save_csv(rows, output_file):
    """
    Сохраняет данные в CSV файл.
    """
    with open(output_file, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(rows)


def save_json_results(results, json_output_file):
    """
    Сохраняет результаты обработки в JSON файл.
    """
    output_data = {
        "metadata": {
            "total_processed": len(results),
            "generated_at": time.strftime("%Y-%m-%d %H:%M:%S")
        },
        "results": results
    }
    
    with open(json_output_file, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)


def save_excel(rows, output_file):
    """
    Сохраняет данные в Excel файл.
    """
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl не установлен")
    
    import openpyxl as _openpyxl  # type: ignore
    wb = _openpyxl.Workbook()
    ws = wb.active
    
    if ws is None:
        raise ValueError("Не удалось создать активный лист")
    
    # Записываем данные
    for row in rows:
        ws.append(row)
    
    # Автоподбор ширины колонок
    for idx, column in enumerate(ws.columns, start=1):
        max_length = 0
        # Получаем букву колонки через индекс
        from openpyxl.utils import get_column_letter  # type: ignore
        column_letter = get_column_letter(idx)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_file)


if __name__ == "__main__":
    # Пути к файлам
    # Скрипт находится в complains_processing, файлы CSV в Downloads
    script_dir = Path(__file__).parent  # complains_processing
    project_root = script_dir.parent  # gaado_backend
    downloads_dir = Path.home() / "Downloads"
    
    # Входной файл из Downloads
    input_file = str(downloads_dir / "So -_ En - List of comms.csv")
    # JSON файл для результатов (сохраняем в папку со скриптом)
    json_output_file = str(script_dir / "So -_ En - List of comms_results.json")
    # Опционально: CSV/Excel файл (можно закомментировать, если не нужен)
    output_file = None  # str(script_dir / "So -_ En - List of comms_processed.csv")
    # Для Excel раскомментируйте следующую строку:
    # output_file = str(script_dir / "So -_ En - List of comms_processed.xlsx")
    
    # Проверяем существование входного файла
    if not Path(input_file).exists():
        print(f"Ошибка: файл {input_file} не найден!")
        exit(1)
    
    # Определяем режим работы (по умолчанию локальный анализ без API)
    use_api = os.environ.get('USE_API', '').lower() in ['true', '1', 'yes']
    
    print("=" * 60)
    if use_api:
        print("Обработка комментариев через Gemini API")
        print("=" * 60)
        
        # Инициализируем клиент заранее
        try:
            client = get_gemini_client()
            print(f"✅ Используется модель: gemini-3-flash-preview")
        except Exception as e:
            print(f"❌ Ошибка инициализации Gemini API: {e}")
            print("\nПроверьте:")
            print("  1. Правильность API ключа")
            print("  2. Доступность интернета")
            print("  3. Доступность модели gemini-3-flash-preview")
            exit(1)
    else:
        print("Обработка комментариев локально (без API)")
        print("=" * 60)
        print("✅ Используется локальный анализ на основе ключевых слов")
    
    # Показываем формат вывода
    print(f"Формат вывода: JSON ({json_output_file})")
    if output_file:
        output_format = "Excel" if Path(output_file).suffix.lower() in ['.xlsx', '.xls'] else "CSV"
        if output_format == "Excel" and not EXCEL_AVAILABLE:
            print("⚠️  Внимание: openpyxl не установлен, будет использован CSV формат")
            print("   Установите: pip install openpyxl")
        else:
            print(f"Дополнительный формат: {output_format} ({output_file})")
    print("")
    
    try:
        process_csv(input_file, output_file, json_output_file, use_api=use_api)
    except KeyboardInterrupt:
        print("\n\nПрервано пользователем")
    except Exception as e:
        print(f"\nОшибка: {e}")
        import traceback
        traceback.print_exc()

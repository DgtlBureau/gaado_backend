#!/usr/bin/env python3
"""
Скрипт для преобразования JSON результатов в Excel файл.
"""

import json
import os
from pathlib import Path

# Попытка импортировать openpyxl для Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  openpyxl не установлен. Установите: pip install openpyxl")
    exit(1)


def json_to_excel(json_file, excel_file):
    """
    Преобразует JSON файл с результатами в Excel файл.
    """
    # Читаем JSON файл
    print(f"Чтение JSON файла: {json_file}")
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    results = data.get('results', [])
    metadata = data.get('metadata', {})
    
    if not results:
        print("⚠️  JSON файл не содержит результатов!")
        return
    
    print(f"Найдено записей: {len(results)}")
    
    # Создаем Excel книгу
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Risk Analysis Results"
    
    # Заголовки колонок
    headers = [
        "Row Number",
        "Somali",
        "English",
        "Risk Category",
        "Risk Subcategory",
        "Risk Level"
    ]
    
    # Стили для заголовков
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Записываем заголовки
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Записываем данные
    print("Запись данных в Excel...")
    for row_idx, result in enumerate(results, start=2):
        ws.cell(row=row_idx, column=1, value=result.get('row_number', ''))
        ws.cell(row=row_idx, column=2, value=result.get('somali', ''))
        ws.cell(row=row_idx, column=3, value=result.get('english', ''))
        ws.cell(row=row_idx, column=4, value=result.get('risk_category', ''))
        ws.cell(row=row_idx, column=5, value=result.get('risk_subcategory', ''))
        ws.cell(row=row_idx, column=6, value=result.get('risk_level', ''))
        
        # Применяем границы к ячейкам
        for col_idx in range(1, 7):
            ws.cell(row=row_idx, column=col_idx).border = border
    
    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 12  # Row Number
    ws.column_dimensions['B'].width = 50  # Somali
    ws.column_dimensions['C'].width = 60  # English
    ws.column_dimensions['D'].width = 25  # Risk Category
    ws.column_dimensions['E'].width = 30  # Risk Subcategory
    ws.column_dimensions['F'].width = 15  # Risk Level
    
    # Выравнивание текста
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            if cell.column == 1:  # Row Number - по центру
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            else:  # Текст - по левому краю с переносом
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Фиксируем первую строку (заголовки)
    ws.freeze_panes = 'A2'
    
    # Добавляем информацию о метаданных на отдельный лист
    if metadata:
        ws_meta = wb.create_sheet("Metadata")
        ws_meta.append(["Metadata"])
        ws_meta.append(["Total Processed", metadata.get('total_processed', 'N/A')])
        ws_meta.append(["Generated At", metadata.get('generated_at', 'N/A')])
        
        # Стили для метаданных
        ws_meta['A1'].font = Font(bold=True, size=14)
        ws_meta.column_dimensions['A'].width = 20
        ws_meta.column_dimensions['B'].width = 30
    
    # Сохраняем файл
    print(f"Сохранение Excel файла: {excel_file}")
    wb.save(excel_file)
    print(f"✅ Excel файл успешно создан: {excel_file}")
    print(f"   Всего строк: {len(results)}")
    print(f"   Листы: {', '.join([sheet.title for sheet in wb.worksheets])}")


if __name__ == "__main__":
    # Пути к файлам
    script_dir = Path(__file__).parent
    json_file = script_dir / "So -_ En - List of comms_results.json"
    excel_file = script_dir / "So -_ En - List of comms_results.xlsx"
    
    # Проверяем существование JSON файла
    if not json_file.exists():
        print(f"❌ Ошибка: файл {json_file} не найден!")
        exit(1)
    
    print("=" * 60)
    print("Преобразование JSON в Excel")
    print("=" * 60)
    print()
    
    try:
        json_to_excel(json_file, excel_file)
    except Exception as e:
        print(f"\n❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()
